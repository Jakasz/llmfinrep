"""Excel (.xlsx) text extraction using openpyxl.

Iterates all sheets and rows, joining cell values as tab-separated text.
"""

import io
import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def extract_text_from_excel(file_bytes: bytes) -> str:
    """Extract text from an Excel file.

    Args:
        file_bytes: Raw .xlsx file content.

    Returns:
        Text representation of all sheets, rows, and cells.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"--- Sheet: {sheet_name} ---"]

        for row in ws.iter_rows(values_only=True):
            # Перетворити кожну комірку в рядок, пропустити None
            cells = [str(cell) if cell is not None else "" for cell in row]
            line = "\t".join(cells)
            # Пропускати повністю порожні рядки
            if line.strip():
                sheet_lines.append(line)

        parts.append("\n".join(sheet_lines))
        logger.debug("Sheet '%s': extracted %d rows", sheet_name, len(sheet_lines) - 1)

    wb.close()

    result = "\n\n".join(parts)
    logger.info("Excel extraction completed: %d sheets, %d characters", len(wb.sheetnames), len(result))
    return result
